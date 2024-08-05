import os
import requests
import openpyxl
from datetime import datetime

# Discourse API configuration
DISCOURSE_BASE_URL = ""
API_KEY = ""
API_USERNAME = ""

# Get the current working directory
current_dir = os.path.dirname(os.path.abspath(__file__))

# Construct the Excel file path
EXCEL_FILE_PATH = os.path.join(current_dir, "articles.xlsx")

def handle_response(response):
    if response.status_code == 200:
        try:
            response_json = response.json()
            print(f"Successfully posted article: {response_json.get('title', 'Title unavailable')}")
        except (ValueError, KeyError):
            print(f"Error parsing response: {response.text}")
    else:
        print(f"Error posting article: {response.text}")

def post_articles_from_excel():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    except FileNotFoundError:
        print(f"Error: Excel file '{EXCEL_FILE_PATH}' not found.")
        return

    sheet = wb.active

    # Find the last column and insert a new column for status
    last_column = sheet.max_column
    sheet.insert_cols(last_column + 1)
    status_column = sheet.cell(row=1, column=last_column + 1)
    status_column.value = "Status"

    for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        title = row[0].value
        content = row[1].value
        topic_id = row[2].value
        category_id = row[3].value

        data = {
            "title": title,
            "raw": content,
            "category": category_id,
            "created_at": datetime.now().isoformat() + "+00:00",
            "topic_id": topic_id
        }

        url = f"{DISCOURSE_BASE_URL}/posts.json"
        headers = {
            "Api-Key": API_KEY,
            "Api-Username": API_USERNAME,
            "Content-Type": "application/json"
        }

        try:
            response = requests.post(url, headers=headers, json=data)
            response.raise_for_status()  # Raise an exception for non-200 status codes
            sheet.cell(row=row_num, column=5).value = "Success"
            handle_response(response)
        except requests.exceptions.RequestException as e:
            sheet.cell(row=row_num, column=5).value = f"Failed: {e}"
            print(f"Error posting article: {e}")

    wb.save(EXCEL_FILE_PATH)

if __name__ == "__main__":
    post_articles_from_excel()
